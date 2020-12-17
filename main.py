
import glob
import openpyxl
from pathlib import Path

def printData(sheet):
    for row in sheet.iter_rows():
        for cell in row:
            print(cell.value, end=" ")
        print()

    print(sheet.max_row, sheet.max_column)

def get_ValueAt(sheet, row_,col_):
    return sheet.cell(row=row_, column=col_).value

def get_DocumentNr(sheet, row_):
    doc_no = get_ValueAt(sheet,row_,1)
    doc_no = doc_no.replace(" ", "")
    doc_no = doc_no.replace(u'\xa0', u' ')
    return doc_no

def get_Rev(sheet, row_):
    rev = get_ValueAt(sheet,row_,3)    
    if rev:
        return int(rev)
    else:
        return -1

def get_SemiNrFor(sheet, row_):

    PT_Rev = get_ValueAt(sheet,row_,5)
    IC_PartNo = get_DocumentNr(sheet, row_)
    # check if the next article (Sorted list) is the same IC_PartNo => get the latest revision    
    i = 1     
    lastRev = PT_Rev   
    while IC_PartNo == get_DocumentNr(sheet, row_ + i):
        i_PT_Rev = get_ValueAt(sheet,row_ + i ,5)
        if ord(i_PT_Rev) > ord(PT_Rev) :
            lastRev = i_PT_Rev
        i += 1

    PT_PartNo = get_ValueAt(sheet,row_,4)
    semiNr = str(PT_PartNo) + '_' + str(lastRev)
    
    return semiNr

def create_SQL_Insert_Statement(item):
    """ Target Statement
		UPDATE test_tables.dm_document d
        INNER join dm_version v on d.doc_did=v.ver_did
        INNER join dm_prop_amicra p on v.ver_vid=p.prop_did
        SET SEMI_NR = '01-YA'
        WHERE ver_minor= 0 AND ver_major = '6' AND doc_docno = '000001-005301-B04';
    """
    join = '''
    INNER join dm_version v on d.doc_did=v.ver_did
    INNER join dm_prop_amicra p on v.ver_vid=p.prop_did 
    '''

    table = 'test_tables.dm_document d'
    value =  "'"  + str(item['SemiNr']) + "'"
    condition = ' v.ver_minor= 0 AND v.ver_major = ' + "'" + str(item['IC_Rev']) + "'"
    condition += ' AND d.doc_docno = ' + "'" + str(item['IC_PartNo']) + "'"

    statement = """
    UPDATE """
    statement += table + join +'SET SEMI_NR = '+ value + ' WHERE ' + condition +' ;'  
    # UPDATE test_tables.dm_prop_amicra SET SEMI_NR = '26-YA00222_A' WHERE ( prop_pid = '24546' and BEN2_ID = '0') ; 
    return statement

def get_MappingData():
    mappingTable_path = Path('X:/Sanchez Cristian/Semi_Nr/Aktuelle_Mapping_Table_ASIEN/AMICRA Mapping  (2020Dec14 last updated).xlsx')

    mappingTable_wbs = openpyxl.load_workbook(mappingTable_path)
    sheet_AMICRA_Mapping = mappingTable_wbs['AMICRA Mapping']

    # Liste aller Artikeln (IC_PartNo) mit unterschiedlichen IC_Rev (Revisionnummer)
    list_ArticlesAndRev = []
    list_ArticlesWithSemiNr = []

    #printData(sheet_AMICRA_Mapping)
    #print(get_ValueAt(sheet_AMICRA_Mapping,2,5))
    #print(get_SemiNrFor(sheet_AMICRA_Mapping,2))

    #for r in range(2,sheet_AMICRA_Mapping.max_row):
    #    for c in range(1,sheet_AMICRA_Mapping.max_column):
    #        print(sheet_AMICRA_Mapping.cell(row=r, column=c).value)

    for r in range(2,sheet_AMICRA_Mapping.max_row):

        IC_PartNo = get_DocumentNr(sheet_AMICRA_Mapping, r)
        IC_Rev = get_Rev(sheet_AMICRA_Mapping,r)

        if IC_PartNo and (IC_Rev != -1):

            semi_Nr = get_SemiNrFor(sheet_AMICRA_Mapping,r)
            article = {'IC_PartNo' : IC_PartNo, 'IC_Rev' : IC_Rev}
            #print(IC_PartNo)
            if not (article in list_ArticlesAndRev):
                list_ArticlesAndRev.append(article)
                list_ArticlesWithSemiNr.append({'IC_PartNo' : IC_PartNo, 'IC_Rev' : IC_Rev, 'SemiNr' : semi_Nr})

        # print(statement)
    #print(list_ArticlesWithSemiNr)
    
    with open('X:/Sanchez Cristian/Semi_Nr/skript/Insert_all_SemiNr.sql', 'w') as writer:
        for item in list_ArticlesWithSemiNr:        
            statement = create_SQL_Insert_Statement(item)
            writer.write(statement)
            writer.write('\n')


if __name__ == "__main__":
    get_MappingData()
